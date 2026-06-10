from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO
from typing import Any

from .._options import Option, register_options
from ..types import ERROR_PARSE, error_artifact, make_artifact, missing_dep_artifact
from . import register_processor

_SHEET_SEPARATOR = "\n\n"


def _csv_escape(value: Any) -> str:
    if value is None:
        s = ""
    else:
        s = str(value)
    if any(c in s for c in [",", "\n", '"']):
        s = '"' + s.replace('"', '""') + '"'
    return s


def _select_sheet(names: list[str], sheet: str | int | None) -> str | None:
    """Resolve the ``sheet`` option to a sheet name; ``None`` means "all sheets".

    Unknown names and out-of-range indexes fall back to the first sheet
    (a single-sheet render), preserving historical behavior.

    Examples:
        >>> _select_sheet(["A", "B"], None) is None
        True
        >>> _select_sheet(["A", "B"], "B")
        'B'
        >>> _select_sheet(["A", "B"], 1)
        'B'
        >>> _select_sheet(["A", "B"], "Nope")
        'A'
        >>> _select_sheet(["A", "B"], 7)
        'A'
    """
    if sheet is None:
        return None
    if isinstance(sheet, str) and sheet in names:
        return sheet
    if (
        isinstance(sheet, int)
        and not isinstance(sheet, bool)
        and 0 <= sheet < len(names)
    ):
        return names[sheet]
    return names[0] if names else None


def _render_sheet_block(
    name: str, rows: Iterable[tuple[Any, ...]], max_rows: int
) -> tuple[str, bool]:
    """Render one sheet as a ``# <name>`` heading plus CSV lines.

    At most ``max_rows`` rows are kept after the first (header) row.
    Returns ``(block_text, truncated)``.

    Examples:
        >>> block, truncated = _render_sheet_block("Sales", [("a", "b"), (1, 2)], 200)
        >>> print(block)
        # Sales
        a,b
        1,2
        >>> truncated
        False
        >>> _render_sheet_block("S", [("h",), (1,), (2,)], 1)
        ('# S\\nh\\n1', True)
    """
    lines = [f"# {name}"]
    truncated = False
    for i, row in enumerate(rows):
        if i > max_rows:
            truncated = True
            break
        lines.append(",".join(_csv_escape(v) for v in row))
    return "\n".join(lines), truncated


def _join_blocks_with_segments(
    blocks: list[tuple[str, str]],
) -> tuple[str, list[dict[str, Any]]]:
    """Join rendered sheet blocks and build ``meta.segments`` (IR: xlsx sheets).

    Each segment slices exactly one sheet's rendered block, heading included.

    Examples:
        >>> text, segs = _join_blocks_with_segments(
        ...     [("A", "# A\\n1"), ("B", "# B\\n2")]
        ... )
        >>> text
        '# A\\n1\\n\\n# B\\n2'
        >>> segs[1]
        {'kind': 'sheet', 'label': 'B', 'start': 7, 'end': 12}
        >>> text[segs[0]["start"] : segs[0]["end"]]
        '# A\\n1'
    """
    parts: list[str] = []
    segments: list[dict[str, Any]] = []
    offset = 0
    for label, block in blocks:
        if parts:
            offset += len(_SHEET_SEPARATOR)
        segments.append(
            {
                "kind": "sheet",
                "label": label,
                "start": offset,
                "end": offset + len(block),
            }
        )
        parts.append(block)
        offset += len(block)
    return _SHEET_SEPARATOR.join(parts), segments


def _xlsx_with_openpyxl(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    """Primary backend. Returns ``(text, segments, extra)``."""
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        chosen = _select_sheet(names, sheet)
        targets = [chosen] if chosen is not None else names

        blocks: list[tuple[str, str]] = []
        truncated = False
        for name in targets:
            block, t = _render_sheet_block(
                name, wb[name].iter_rows(values_only=True), max_rows
            )
            truncated = truncated or t
            blocks.append((name, block))

        text, segments = _join_blocks_with_segments(blocks)
        extra: dict[str, Any] = {"sheets": names, "engine": "openpyxl"}
        if len(targets) == 1:
            ws = wb[targets[0]]
            extra["sheet_used"] = targets[0]
            extra["rows"] = ws.max_row
            extra["cols"] = ws.max_column
        if truncated:
            extra["rows_truncated"] = True
        return text, segments, extra
    finally:
        wb.close()


def _xlsx_with_pandas(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    """Fallback backend; renders the same text layout as the openpyxl path."""
    import pandas as pd  # type: ignore

    xls = pd.ExcelFile(BytesIO(data))
    names: list[str] = list(xls.sheet_names)
    chosen = _select_sheet(names, sheet)
    targets = [chosen] if chosen is not None else names

    blocks: list[tuple[str, str]] = []
    truncated = False
    shape: tuple[int, int] | None = None
    for name in targets:
        # header=None keeps the header as a plain first row, matching the
        # openpyxl path; NaN cells become None so both render as "".
        df = xls.parse(name, header=None)
        df = df.astype(object).where(pd.notna(df), None)
        block, t = _render_sheet_block(
            name, df.itertuples(index=False, name=None), max_rows
        )
        truncated = truncated or t
        blocks.append((name, block))
        shape = (int(df.shape[0]), int(df.shape[1]))

    text, segments = _join_blocks_with_segments(blocks)
    extra: dict[str, Any] = {"sheets": names, "engine": "pandas"}
    if len(targets) == 1 and shape is not None:
        extra["sheet_used"] = targets[0]
        extra["rows"], extra["cols"] = shape
    if truncated:
        extra["rows_truncated"] = True
    return text, segments, extra


def _normalize_xls_value(value: Any) -> Any:
    """Normalize an xlrd cell value for layout parity with openpyxl.

    xlrd returns ALL numbers as floats; map integral floats to int so a
    cell that stores ``25`` renders as ``"25"`` rather than ``"25.0"``.
    Empty cells arrive as ``''`` and already render as ``''``. Dates render
    as their raw serial floats (documented, accepted for 1.0).

    Examples:
        >>> _normalize_xls_value(25.0)
        25
        >>> _normalize_xls_value(2.5)
        2.5
        >>> _normalize_xls_value("")
        ''
        >>> _normalize_xls_value("hi")
        'hi'
    """
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _xls_with_xlrd(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    """Legacy .xls backend (xlrd). Returns ``(text, segments, extra)``."""
    import xlrd  # type: ignore

    wb = xlrd.open_workbook(file_contents=data)
    names: list[str] = list(wb.sheet_names())
    chosen = _select_sheet(names, sheet)
    targets = [chosen] if chosen is not None else names

    blocks: list[tuple[str, str]] = []
    truncated = False
    for name in targets:
        sh = wb.sheet_by_name(name)
        rows = (
            tuple(_normalize_xls_value(v) for v in sh.row_values(r))
            for r in range(sh.nrows)
        )
        block, t = _render_sheet_block(name, rows, max_rows)
        truncated = truncated or t
        blocks.append((name, block))

    text, segments = _join_blocks_with_segments(blocks)
    extra: dict[str, Any] = {"sheets": names, "engine": "xlrd"}
    if len(targets) == 1:
        sh = wb.sheet_by_name(targets[0])
        extra["sheet_used"] = targets[0]
        extra["rows"] = sh.nrows
        extra["cols"] = sh.ncols
    if truncated:
        extra["rows_truncated"] = True
    return text, segments, extra


def xls_processor(data: bytes, **options: Any) -> dict[str, Any]:
    """Legacy Excel (.xls, BIFF) -> artifact via xlrd. Same layout, options,
    and ``meta.segments`` semantics as :func:`xlsx_processor`: all sheets by
    default, each a ``# <sheet name>`` heading plus CSV rows.
    """
    sheet = options.get("sheet")
    max_rows = int(options.get("max_rows", 200))
    source = options.get("filename") or "workbook.xls"

    try:
        import xlrd  # type: ignore # noqa: F401
    except ImportError:
        return missing_dep_artifact(source, "xls")

    try:
        text, segments, extra = _xls_with_xlrd(data, sheet=sheet, max_rows=max_rows)
    except Exception as e:
        return error_artifact(
            source, ERROR_PARSE, f"Failed to parse legacy Excel workbook: {e}"
        )

    meta: dict[str, Any] = {"kind": "table", "extra": extra}
    if segments:
        meta["segments"] = segments
    return make_artifact(text=text, meta=meta)


def xlsx_processor(data: bytes, **options: Any) -> dict[str, Any]:
    """XLSX -> artifact. No ``sheet`` option: render ALL sheets, each as a
    ``# <sheet name>`` heading plus CSV rows, joined with a blank line, with
    one ``meta.segments`` entry per sheet. With ``sheet``: render just that
    sheet (still with its heading and single segment).
    """
    sheet = options.get("sheet")
    max_rows = int(options.get("max_rows", 200))
    source = options.get("filename") or "workbook.xlsx"

    openpyxl_exc = None
    pandas_exc = None

    # Primary: openpyxl (segments come straight from the rendered blocks).
    try:
        text, segments, extra = _xlsx_with_openpyxl(
            data, sheet=sheet, max_rows=max_rows
        )
        meta: dict[str, Any] = {"kind": "table", "extra": extra}
        if segments:
            meta["segments"] = segments
        return make_artifact(text=text, meta=meta)
    except ImportError:  # pragma: no cover - optional dep
        pass
    except Exception as e:
        openpyxl_exc = str(e)

    # Fallback: pandas, rendering the identical text layout.
    try:
        text, segments, extra = _xlsx_with_pandas(data, sheet=sheet, max_rows=max_rows)
        meta = {"kind": "table", "extra": extra}
        if segments:
            meta["segments"] = segments
        return make_artifact(text=text, meta=meta)
    except ImportError:  # pragma: no cover - optional dep
        pass
    except Exception as e:
        pandas_exc = str(e)

    # Neither openpyxl nor pandas available
    if openpyxl_exc is None and pandas_exc is None:
        return missing_dep_artifact(source, "xlsx")

    # Backends were importable but the workbook could not be parsed
    detail = openpyxl_exc or pandas_exc
    artifact = error_artifact(
        source, ERROR_PARSE, f"Failed to parse Excel workbook: {detail}"
    )
    artifact["meta"]["extra"] = {
        "pandas_exc": pandas_exc,
        "openpyxl_exc": openpyxl_exc,
    }
    return artifact


_SHEET_OPTIONS = (
    Option(
        "sheet",
        "str_or_int",
        help="Sheet to render: a sheet name or 0-based index. "
        "Omit to render all sheets.",
        example="sheet: Sales",
    ),
    Option(
        "rows",
        "int",
        aliases=("max_rows",),
        param="max_rows",
        default=200,
        help="Maximum number of rows rendered as text per sheet.",
        example="rows: 100",
    ),
)

register_processor(".xlsx", xlsx_processor)
register_options(".xlsx", _SHEET_OPTIONS)

register_processor(".xls", xls_processor)
register_options(".xls", _SHEET_OPTIONS)
