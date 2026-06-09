from __future__ import annotations

from io import BytesIO
from typing import Any

from ..types import ERROR_PARSE, error_artifact, make_artifact, missing_dep_artifact
from . import register_processor


def _sheet_segments(text: str, sheet_name: str) -> list[dict[str, Any]]:
    """Build meta.segments for the rendered sheet (IR contract: xlsx sheets).

    Both backends render a single chosen sheet, so the segmentation is one
    sheet segment spanning the whole rendered text.

    Examples:
        >>> _sheet_segments("a,b\\n1,2", "Sales")
        [{'kind': 'sheet', 'label': 'Sales', 'start': 0, 'end': 7}]
    """
    return [{"kind": "sheet", "label": sheet_name, "start": 0, "end": len(text)}]


def _csv_escape(value: Any) -> str:
    if value is None:
        s = ""
    else:
        s = str(value)
    if any(c in s for c in [",", "\n", '"']):
        s = '"' + s.replace('"', '""') + '"'
    return s


def _xlsx_with_pandas(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, dict[str, Any]]:
    import pandas as pd  # type: ignore

    # Discover sheets
    xls = pd.ExcelFile(BytesIO(data))
    sheet_names: list[str] = list(xls.sheet_names)
    # Choose sheet
    chosen: str
    if isinstance(sheet, str) and sheet in sheet_names:
        chosen = sheet
    elif isinstance(sheet, int) and 0 <= sheet < len(sheet_names):
        chosen = sheet_names[sheet]
    else:
        chosen = sheet_names[0] if sheet_names else "Sheet1"

    df = xls.parse(chosen)
    # Render as CSV text for broad compatibility
    head = df.head(max_rows)
    text = head.to_csv(index=False)
    extra = {
        "rows": int(df.shape[0]),
        "cols": int(df.shape[1]),
        "sheets": sheet_names,
        "sheet_used": chosen,
        "engine": "pandas",
    }
    return text, extra


def _xlsx_with_openpyxl(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    names = list(wb.sheetnames)

    if isinstance(sheet, str) and sheet in names:
        chosen = sheet
    elif isinstance(sheet, int) and 0 <= sheet < len(names):
        chosen = names[sheet]
    else:
        chosen = names[0] if names else "Sheet1"

    ws = wb[chosen]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > max_rows:
            break
        rows.append([_csv_escape(v) for v in row])
    text = "\n".join([",".join(r) for r in rows])
    extra = {
        "rows": ws.max_row,
        "cols": ws.max_column,
        "sheets": names,
        "sheet_used": chosen,
        "engine": "openpyxl",
    }
    return text, extra


def xlsx_processor(data: bytes, **options: Any) -> dict[str, Any]:
    sheet = options.get("sheet")
    max_rows = int(options.get("max_rows", 200))
    source = options.get("filename") or "workbook.xlsx"

    pandas_exc = None
    openpyxl_exc = None

    # Prefer pandas if available
    try:
        text, extra = _xlsx_with_pandas(data, sheet=sheet, max_rows=max_rows)
        return make_artifact(
            text=text,
            meta={
                "kind": "table",
                "extra": extra,
                "segments": _sheet_segments(text, extra["sheet_used"]),
            },
        )
    except ImportError:  # pragma: no cover - optional dep
        pass
    except Exception as e:
        pandas_exc = str(e)

    # Fallback to openpyxl-only
    try:
        text, extra = _xlsx_with_openpyxl(data, sheet=sheet, max_rows=max_rows)
        return make_artifact(
            text=text,
            meta={
                "kind": "table",
                "extra": extra,
                "segments": _sheet_segments(text, extra["sheet_used"]),
            },
        )
    except ImportError:  # pragma: no cover - optional dep
        pass
    except Exception as e:
        openpyxl_exc = str(e)

    # Neither pandas nor openpyxl available
    if pandas_exc is None and openpyxl_exc is None:
        return missing_dep_artifact(source, "xlsx")

    # Backends were importable but the workbook could not be parsed
    detail = pandas_exc or openpyxl_exc
    artifact = error_artifact(
        source, ERROR_PARSE, f"Failed to parse Excel workbook: {detail}"
    )
    artifact["meta"]["extra"] = {
        "pandas_exc": pandas_exc,
        "openpyxl_exc": openpyxl_exc,
    }
    return artifact


register_processor(".xlsx", xlsx_processor)
