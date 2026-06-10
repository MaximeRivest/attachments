"""
Tests for the Excel (XLSX) processor.

=============================================================================
TEST GUIDELINES FOR XLSX PROCESSOR
=============================================================================

GOOD tests for XLSX processor:
    - Use pytest.mark.skipif for optional deps
    - Test with real XLSX bytes (fixture from conftest.py or built in-test)
    - Test the all-sheets default and sheet selection options
    - Test row limiting per sheet
    - Test fallback between openpyxl and pandas

BAD tests:
    - Assuming pandas/openpyxl are always installed
    - Huge test files (use minimal fixtures)

NOTES:
    - XLSX processor uses openpyxl as the primary backend, pandas as fallback
    - Output is a "# <sheet name>" heading + CSV rows per sheet; with no
      'sheet' option ALL sheets are rendered, joined with a blank line
    - meta.segments carries one sheet segment per rendered sheet
    - Tests skip gracefully if deps not installed

=============================================================================
"""

from __future__ import annotations

from io import BytesIO

import pytest

from attachments._processors import processors
from attachments.deps import check_dep

# Skip all tests if no XLSX deps available
pytestmark = pytest.mark.skipif(
    not check_dep("xlsx").available,
    reason="XLSX deps (openpyxl) not installed",
)


def _build_workbook(sheets: dict[str, list[list]]) -> bytes:
    """Build a real multi-sheet XLSX in memory with openpyxl."""
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def multi_sheet_xlsx_bytes() -> bytes:
    """Three-sheet workbook for segmentation tests."""
    return _build_workbook(
        {
            "Alpha": [["Name", "Age"], ["Alice", 30], ["Bob", 25]],
            "Beta": [["Product", "Revenue"], ["Widget", 1000]],
            "Gamma": [["X"], [1], [2], [3]],
        }
    )


class TestXlsxProcessor:
    """Tests for XLSX processor with real Excel files."""

    def test_processor_registered(self):
        assert ".xlsx" in processors

    def test_returns_artifact_structure(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        assert "text" in result
        assert "images" in result
        assert "audio" in result
        assert "video" in result
        assert "meta" in result

    def test_extracts_text_as_csv(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        # Should contain header row
        assert "Name" in result["text"]
        assert "Age" in result["text"]
        assert "City" in result["text"]

        # Should contain data
        assert "Alice" in result["text"]
        assert "30" in result["text"]

    def test_all_sheets_rendered_by_default(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        # Both sheets present, each with a heading line
        assert "# Sheet1" in result["text"]
        assert "# Sales" in result["text"]
        assert "Product" in result["text"]
        assert "Widget" in result["text"]
        # No sheet_used when several sheets are rendered
        assert "sheet_used" not in result["meta"]["extra"]

    def test_meta_includes_metadata(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        meta = result["meta"]
        assert meta["kind"] == "table"
        extra = meta["extra"]
        assert "sheets" in extra
        assert "engine" in extra

    def test_single_sheet_metadata(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet="Sheet1")

        extra = result["meta"]["extra"]
        assert extra["sheet_used"] == "Sheet1"
        assert "rows" in extra
        assert "cols" in extra

    def test_extra_engine_is_pandas_or_openpyxl(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        assert result["meta"]["extra"]["engine"] in ("pandas", "openpyxl")


class TestXlsxProcessorOptions:
    """Tests for XLSX processor options."""

    def test_sheet_selection_by_name(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet="Sales")

        # Should use the Sales sheet only
        assert result["meta"]["extra"]["sheet_used"] == "Sales"
        assert result["text"].startswith("# Sales\n")
        assert "Product" in result["text"]
        assert "Revenue" in result["text"]
        assert "# Sheet1" not in result["text"]

    def test_sheet_selection_by_index(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet=1)  # Second sheet (Sales)

        assert result["meta"]["extra"]["sheet_used"] == "Sales"
        assert result["text"].startswith("# Sales\n")

    def test_sheet_selection_invalid_uses_first(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet="NonexistentSheet")

        # Should fall back to first sheet
        assert result["meta"]["extra"]["sheet_used"] == "Sheet1"

    def test_max_rows_limits_output_per_sheet(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes, max_rows=1)

        # Each sheet block: heading + header row + at most 1 data row
        for segment in result["meta"]["segments"]:
            block = result["text"][segment["start"] : segment["end"]]
            assert len(block.split("\n")) <= 3

        # Alpha (2 data rows) and Gamma (3 data rows) were truncated
        assert result["meta"]["extra"]["rows_truncated"] is True

    def test_max_rows_with_single_sheet(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet="Sheet1", max_rows=2)

        # heading + header + 2 data rows
        lines = result["text"].strip().split("\n")
        assert len(lines) == 4
        assert result["meta"]["extra"]["rows_truncated"] is True

    def test_rows_truncated_absent_when_not_hit(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        # Default 200 rows is plenty for the fixture
        assert "rows_truncated" not in result["meta"]["extra"]

    def test_default_max_rows(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        # Default is 200 rows - our test file is smaller
        assert "text" in result


class TestXlsxProcessorErrors:
    """Tests for XLSX processor error handling."""

    def test_corrupt_xlsx_returns_parse_error(self, corrupt_xlsx_bytes: bytes):
        from attachments.types import ERROR_PARSE

        processor = processors[".xlsx"]
        result = processor(corrupt_xlsx_bytes)

        # Should return artifact with a typed error, not raise
        assert result["text"] == ""
        assert result["meta"]["error"]["code"] == ERROR_PARSE

    def test_empty_bytes_handled(self):
        processor = processors[".xlsx"]
        result = processor(b"")

        assert "meta" in result


class TestXlsxSheetDiscovery:
    """Tests for sheet discovery."""

    def test_sheets_list_populated(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        sheets = result["meta"]["extra"]["sheets"]
        assert isinstance(sheets, list)
        assert "Sheet1" in sheets
        assert "Sales" in sheets

    def test_sheet_count(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes)

        assert len(result["meta"]["extra"]["sheets"]) == 2


@pytest.mark.skipif(
    not check_dep("xlsx-pandas").available,
    reason="pandas not installed",
)
class TestXlsxWithPandas:
    """The pandas fallback renders the same text layout as openpyxl."""

    def test_pandas_layout_matches_openpyxl(self, multi_sheet_xlsx_bytes: bytes):
        from attachments._processors.xlsx import (
            _xlsx_with_openpyxl,
            _xlsx_with_pandas,
        )

        text_o, segs_o, _ = _xlsx_with_openpyxl(
            multi_sheet_xlsx_bytes, sheet=None, max_rows=200
        )
        text_p, segs_p, extra_p = _xlsx_with_pandas(
            multi_sheet_xlsx_bytes, sheet=None, max_rows=200
        )

        assert text_p == text_o
        assert segs_p == segs_o
        assert extra_p["engine"] == "pandas"

    def test_pandas_single_sheet_matches_openpyxl(self, sample_xlsx_bytes: bytes):
        from attachments._processors.xlsx import (
            _xlsx_with_openpyxl,
            _xlsx_with_pandas,
        )

        text_o, segs_o, _ = _xlsx_with_openpyxl(
            sample_xlsx_bytes, sheet="Sales", max_rows=200
        )
        text_p, segs_p, extra_p = _xlsx_with_pandas(
            sample_xlsx_bytes, sheet="Sales", max_rows=200
        )

        assert text_p == text_o
        assert segs_p == segs_o
        assert extra_p["sheet_used"] == "Sales"


class TestXlsxSegments:
    """meta.segments carries one sheet segment per rendered sheet (IR contract)."""

    def test_all_sheets_segments_slice_exactly(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes)

        text = result["text"]
        segments = result["meta"]["segments"]
        names = result["meta"]["extra"]["sheets"]
        assert names == ["Alpha", "Beta", "Gamma"]
        assert [s["label"] for s in segments] == names

        # Each segment slices exactly that sheet's block, heading included
        blocks = text.split("\n\n")
        assert len(blocks) == 3
        for segment, name, block in zip(segments, names, blocks, strict=False):
            assert segment["kind"] == "sheet"
            assert text[segment["start"] : segment["end"]] == block
            assert block.startswith(f"# {name}\n")

        # Segments tile the text: start at 0, end at len(text), separated
        # by exactly the two-character joiner.
        assert segments[0]["start"] == 0
        assert segments[-1]["end"] == len(text)
        for prev, nxt in zip(segments, segments[1:], strict=False):
            assert nxt["start"] == prev["end"] + 2

    def test_all_sheets_segment_content(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes)

        text = result["text"]
        by_label = {s["label"]: s for s in result["meta"]["segments"]}
        beta = text[by_label["Beta"]["start"] : by_label["Beta"]["end"]]
        assert beta == "# Beta\nProduct,Revenue\nWidget,1000"

    def test_single_sheet_still_has_segment(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes, sheet="Beta")

        segments = result["meta"]["segments"]
        assert len(segments) == 1
        segment = segments[0]
        assert segment["kind"] == "sheet"
        assert segment["label"] == "Beta"
        assert segment["start"] == 0
        assert segment["end"] == len(result["text"])

    def test_single_sheet_by_index_segment(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes, sheet=2)

        segments = result["meta"]["segments"]
        assert len(segments) == 1
        assert segments[0]["label"] == "Gamma"
        assert result["text"].startswith("# Gamma\n")

    def test_sheet_segment_label_follows_selection(self, sample_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(sample_xlsx_bytes, sheet="Sales")

        assert result["meta"]["segments"][0]["label"] == "Sales"

    def test_segments_respect_max_rows(self, multi_sheet_xlsx_bytes: bytes):
        processor = processors[".xlsx"]
        result = processor(multi_sheet_xlsx_bytes, max_rows=1)

        text = result["text"]
        by_label = {s["label"]: s for s in result["meta"]["segments"]}
        gamma = text[by_label["Gamma"]["start"] : by_label["Gamma"]["end"]]
        # heading + header + exactly 1 data row
        assert gamma == "# Gamma\nX\n1"


# Missing-dependency behavior is covered by the always-runnable tests in
# tests/test_processors/test_missing_deps.py (this module is skipped
# entirely when XLSX deps are absent, so such tests could never run here).
# EXCEPTION: the .xls processor's missing-dep test lives below (this shared
# file must not be edited per package ownership), masking xlrd directly.


# ---------------------------------------------------------------------------
# Legacy Excel (.xls) processor — xlrd backend
# ---------------------------------------------------------------------------
#
# Embedded 2-sheet legacy BIFF .xls fixture (zlib-compressed, base64).
# Generated once with xlwt; reproduce with:
#
#     import xlwt, base64, zlib
#     from io import BytesIO
#     wb = xlwt.Workbook()
#     s1 = wb.add_sheet("Sheet1")
#     for r, row in enumerate(
#         [["Name", "Age"], ["Alice", 30], ["Bob", 25], ["Cara", 41]]
#     ):
#         for c, v in enumerate(row):
#             s1.write(r, c, v)
#     s2 = wb.add_sheet("Sheet2")
#     for r, row in enumerate(
#         [["Product", "Revenue"], ["Widget", 1000], ["Gadget", 2.5]]
#     ):
#         for c, v in enumerate(row):
#             s2.write(r, c, v)
#     buf = BytesIO(); wb.save(buf)
#     print(base64.b64encode(zlib.compress(buf.getvalue(), 9)).decode())

_XLS_FIXTURE_B64 = (
    "eNrtWE1oE0EU/mY3/7RpUlOhFUooWLX2Ur14qTEVzckQ66Eigm6btYbGREIq2INWa46C4En"
    "xUimCl6oXf9CD3jwIFT0IgpDqsaeCgocm65uXCSYaaAMatOwX5s3bN9/87L43byf7dilYnH"
    "/Us4xfsB86ypYXrhqboOKtXgRA7ZYl1WrtoWLZ+K/g9ZAjXU48b3/jlj6U/l6GhoeOVySBz"
    "1RO4jzi2YwZbiFGeA2GkGsYFg6KPQ13qPjRzevqZDnBcgvLB8x9wfIAW66zHCZuUZzAUiQ+"
    "sE/F8XGtj9v8JAWecJ+PbBlCF17LOL58Q1S4TkRzKSP9bzb0OtqwAPJczMyYOSNdRIhcuIB"
    "vVhj4Wt2rL8O2vbV2AbJ/r7e7G9hvag5gFtZpIQOxQAHpdVS24bGzppkfkpaSs8ayZw0JTs"
    "KyEDVunDN1IDppEiuaTk3Iq5HsODUdNHIGRUYil01OT+RJGzUvmJlpk7b4WCo5aeZJiRlS8"
    "cnxOBEE6hJBO2+PNpJJdLAe5E0SoFS/dn/13ZHxROQUW2Y5+VdeEdvlPcHCFdmDOvu5hW+L"
    "uQMsd7O8yqNuY72HZYhCm+r+RJdSDs8x5xq39tM8exnvIztq9J2kF1aOPu0tfInsIn0xtjw"
    "TWvwQmUcfPakk9Ze/OQyKQXH7lsSzSLUWKpl8Ytn9W2LxaAG1dku9BztQgo/VIMvKlXw6Qv"
    "FFA75gvhzxkubj9QQxo/haA77GfF3xK0//jOLrDfg68x2KrzP/nnyVa514LAejtPgTrfC5v"
    "kl97qzzuWtdn7vrfH63Yz2fe+p8vqIrH6KRD23YsGHDhg0bNpqEUAc0XR3RneoY6FbfdUpU"
    "yvZnkk2LUWTpl6e/pYeQoTqHi03Fz1Y4RXUsscE+1e+FEmM0ew5TGOd1TDUdv3SMF7X3s+G"
    "OgT+3hZqdv9zMOv/y/D8A/yQLlA=="
)


def _xls_fixture_bytes() -> bytes:
    import base64
    import zlib

    return zlib.decompress(base64.b64decode(_XLS_FIXTURE_B64))


_needs_xlrd = pytest.mark.skipif(
    not check_dep("xls").available,
    reason="XLS deps (xlrd) not installed",
)


class TestXlsRegistration:
    """Registration and option-schema tests (no xlrd needed)."""

    def test_processor_registered(self):
        from attachments._processors.xlsx import xls_processor

        assert processors[".xls"] is xls_processor

    def test_options_schema_identical_to_xlsx(self):
        from attachments._options import get_options

        assert get_options(".xls") == get_options(".xlsx")


@_needs_xlrd
class TestXlsProcessor:
    """Behavioral tests against the embedded 2-sheet .xls fixture."""

    def test_all_sheets_text_layout(self):
        result = processors[".xls"](_xls_fixture_bytes())

        text = result["text"]
        assert "# Sheet1" in text
        assert "# Sheet2" in text
        assert "Name,Age" in text
        assert "Product,Revenue" in text
        assert result["meta"]["kind"] == "table"
        assert "sheet_used" not in result["meta"]["extra"]

    def test_segments_slice_exactly(self):
        result = processors[".xls"](_xls_fixture_bytes())

        text = result["text"]
        segments = result["meta"]["segments"]
        assert [s["label"] for s in segments] == ["Sheet1", "Sheet2"]
        blocks = text.split("\n\n")
        for segment, block in zip(segments, blocks, strict=True):
            assert segment["kind"] == "sheet"
            assert text[segment["start"] : segment["end"]] == block
        assert segments[0]["start"] == 0
        assert segments[-1]["end"] == len(text)

    def test_sheet_by_name(self):
        result = processors[".xls"](_xls_fixture_bytes(), sheet="Sheet2")

        assert result["text"].startswith("# Sheet2\n")
        assert "# Sheet1" not in result["text"]
        extra = result["meta"]["extra"]
        assert extra["sheet_used"] == "Sheet2"
        assert extra["rows"] == 3
        assert extra["cols"] == 2

    def test_sheet_by_index(self):
        result = processors[".xls"](_xls_fixture_bytes(), sheet=1)

        assert result["meta"]["extra"]["sheet_used"] == "Sheet2"
        assert result["text"].startswith("# Sheet2\n")

    def test_unknown_sheet_name_falls_back_to_first(self):
        result = processors[".xls"](_xls_fixture_bytes(), sheet="Nope")

        assert result["meta"]["extra"]["sheet_used"] == "Sheet1"
        assert result["text"].startswith("# Sheet1\n")

    def test_rows_truncation(self):
        result = processors[".xls"](_xls_fixture_bytes(), max_rows=1)

        text = result["text"]
        for segment in result["meta"]["segments"]:
            block = text[segment["start"] : segment["end"]]
            # heading + header + at most 1 data row
            assert len(block.split("\n")) <= 3
        assert result["meta"]["extra"]["rows_truncated"] is True

    def test_integer_floats_render_without_decimal(self):
        result = processors[".xls"](_xls_fixture_bytes())

        text = result["text"]
        assert "Alice,30" in text
        assert "Bob,25" in text
        assert "Widget,1000" in text
        assert "25.0" not in text
        # Genuine non-integral floats keep their decimal part
        assert "Gadget,2.5" in text

    def test_extra_engine_and_sheets(self):
        result = processors[".xls"](_xls_fixture_bytes())

        extra = result["meta"]["extra"]
        assert extra["engine"] == "xlrd"
        assert extra["sheets"] == ["Sheet1", "Sheet2"]


class TestXlsErrors:
    """Error paths never raise (no xlrd needed for the missing-dep case)."""

    def test_missing_xlrd_returns_typed_error(self, monkeypatch):
        import sys

        from attachments.deps import clear_cache
        from attachments.types import (
            ERROR_MISSING_DEPENDENCY,
            is_missing_dependency,
        )

        monkeypatch.setitem(sys.modules, "xlrd", None)
        clear_cache()
        try:
            result = processors[".xls"](_xls_fixture_bytes())
        finally:
            clear_cache()

        error = result["meta"]["error"]
        assert error["code"] == ERROR_MISSING_DEPENDENCY
        assert "pip install attachments[xls]" in error["message"]
        assert is_missing_dependency(result)
        assert result["text"] == ""

    @_needs_xlrd
    def test_corrupt_bytes_return_parse_error(self):
        from attachments.types import ERROR_PARSE

        result = processors[".xls"](b"this is not a BIFF workbook")

        assert result["text"] == ""
        assert result["meta"]["error"]["code"] == ERROR_PARSE
        assert "legacy Excel" in result["meta"]["error"]["message"]

    @_needs_xlrd
    def test_xlsx_bytes_as_xls_return_parse_error(self, sample_xlsx_bytes: bytes):
        from attachments.types import ERROR_PARSE

        result = processors[".xls"](sample_xlsx_bytes)

        assert result["text"] == ""
        assert result["meta"]["error"]["code"] == ERROR_PARSE

    @_needs_xlrd
    def test_empty_bytes_handled(self):
        result = processors[".xls"](b"")

        assert "meta" in result


@pytest.mark.skipif(
    not (check_dep("xls").available and check_dep("xlsx").available),
    reason="xlrd or openpyxl not installed",
)
class TestXlsXlsxParity:
    """The same logical table renders with the same layout shape in both."""

    def test_layout_matches_xlsx_for_same_table(self):
        from attachments._processors.xlsx import (
            _xls_with_xlrd,
            _xlsx_with_openpyxl,
        )

        # Same logical content as the .xls fixture, written as .xlsx
        xlsx_bytes = _build_workbook(
            {
                "Sheet1": [["Name", "Age"], ["Alice", 30], ["Bob", 25], ["Cara", 41]],
                "Sheet2": [["Product", "Revenue"], ["Widget", 1000], ["Gadget", 2.5]],
            }
        )

        text_x, segs_x, _ = _xlsx_with_openpyxl(xlsx_bytes, sheet=None, max_rows=200)
        text_l, segs_l, extra_l = _xls_with_xlrd(
            _xls_fixture_bytes(), sheet=None, max_rows=200
        )

        assert text_l == text_x
        assert segs_l == segs_x
        assert extra_l["engine"] == "xlrd"
